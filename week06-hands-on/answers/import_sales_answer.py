"""
Step 1 정답: data/incoming/ 의 법인 엑셀 8개 → monthly_sales UPSERT
같은 (corp_code, month) 조합은 덮어쓰기 — 멱등성 보장.

구현 방식: SQLite의 INSERT ... ON CONFLICT DO UPDATE (현업 표준 UPSERT).
- 원자적 (race condition 없음)
- 행마다 SELECT 한 번 더 안 함
- 멱등성 시연은 카운트가 아니라 "총 행 수가 그대로다"로 보여줌
"""
import os
import re
import sqlite3
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.normpath(os.path.join(HERE, "..", "..", "week05-hands-on", "data", "sales.db"))
INCOMING_DIR = os.path.normpath(os.path.join(HERE, "..", "data", "incoming"))

FILENAME_RE = re.compile(r"^법인_([A-Z]{2}\d{2})_.+\.xlsx$")


def ensure_unique_index(conn: sqlite3.Connection) -> None:
    """monthly_sales에 (corp_code, month) UNIQUE 인덱스가 없으면 생성."""
    conn.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_sales_corp_month "
        "ON monthly_sales(corp_code, month)"
    )


def parse_corp_code(filename: str) -> str | None:
    """'법인_US01_미국.xlsx' → 'US01'"""
    m = FILENAME_RE.match(filename)
    return m.group(1) if m else None


def read_excel_rows(path: str) -> list[tuple[str, float, str]]:
    """엑셀에서 (월, 금액, 비고) 행 리스트 반환. 헤더 1행 건너뜀."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row is None or row[0] is None:
            continue
        month = str(row[0]).strip()
        amount = float(row[3]) if row[3] is not None else 0.0
        note = str(row[4]) if row[4] is not None else ""
        rows.append((month, amount, note))
    return rows


def main() -> dict:
    """메인 로직. run_pipeline.py에서 import해 호출 가능.

    Returns: {"files": int, "processed": int, "skipped": int, "total_rows": int}
    """
    if not os.path.exists(DB_PATH):
        raise FileNotFoundError(
            f"DB not found: {DB_PATH}\n"
            "week05-hands-on/data/create_db.py 를 먼저 실행하세요."
        )
    if not os.path.isdir(INCOMING_DIR):
        raise FileNotFoundError(f"Incoming dir not found: {INCOMING_DIR}")

    conn = sqlite3.connect(DB_PATH)
    ensure_unique_index(conn)

    # 유효한 corp_code 셋 (corporations 테이블)
    valid_corps = {
        row[0] for row in conn.execute("SELECT corp_code FROM corporations")
    }

    files_processed = 0
    processed_rows = 0
    skipped = 0

    for filename in sorted(os.listdir(INCOMING_DIR)):
        if not filename.endswith(".xlsx"):
            continue
        if filename.startswith("~$"):  # Excel 임시 잠금 파일
            continue
        corp_code = parse_corp_code(filename)
        if corp_code is None:
            print(f"  [SKIP] 파일명 규칙 불일치: {filename}")
            continue
        if corp_code not in valid_corps:
            print(f"  [SKIP] {filename}: corp_code={corp_code}가 corporations에 없음")
            skipped += 1
            continue

        path = os.path.join(INCOMING_DIR, filename)
        rows = read_excel_rows(path)

        for month, amount, note in rows:
            # 현업 표준 UPSERT — 원자적, race condition 없음
            conn.execute(
                "INSERT INTO monthly_sales (corp_code, month, amount, note) "
                "VALUES (?, ?, ?, ?) "
                "ON CONFLICT(corp_code, month) DO UPDATE SET "
                "  amount = excluded.amount, "
                "  note   = excluded.note",
                (corp_code, month, amount, note),
            )
            processed_rows += 1

        files_processed += 1

    conn.commit()

    # 최종 행 수 확인 — 멱등성은 "총 행 수가 그대로다"로 시연
    total_rows = conn.execute("SELECT COUNT(*) FROM monthly_sales").fetchone()[0]
    conn.close()

    result = {
        "files": files_processed,
        "processed": processed_rows,
        "skipped": skipped,
        "total_rows": total_rows,
    }

    print(f"  처리 파일: {files_processed}개")
    print(f"  처리 행:   {processed_rows}건 (UPSERT)")
    print(f"  SKIPPED:  {skipped}건")
    print(f"  monthly_sales 총 행 수: {total_rows}")

    return result


if __name__ == "__main__":
    main()
