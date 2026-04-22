"""
이상거래 Mock SAP 출고 데이터 생성 (F2 데모: 이상거래처 자동 알림 메일)

영업관리 (이상출고 모니터링) 업무 맥락:
- SAP ZSDR0026류의 출고/반품 raw 데이터
- Fixture(임플란트 부품) 동일 품목코드 50개 이상 출고 + 반품률 0% = 가공매출 의심
- 이상 거래처 자동 탐지 → 담당자 이메일로 경각심 메일 발송

시연 흐름:
  1. 이 xlsx를 Claude에게 주기
  2. "이상 거래처 추출 + 담당자별 개인화 메일 생성 + SMTP로 내 부계정 앞으로 발송"
  3. 부계정 받은편지함에서 실제 메일 확인

실행:
  py week07-hands-on/prep/make_abnormal_shipments.py

출력:
  week07-hands-on/demos/email_send/data/shipments.xlsx
"""

from __future__ import annotations

import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")  # cp949 → utf-8 (한글/대시 안전 출력)


HERE = Path(__file__).parent
OUT_DIR = HERE.parent / "demos" / "email_send" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / "shipments.xlsx"


# 품목 마스터 (Fixture: 치과 임플란트 부품)
ITEMS = {
    "FX001": ("임플란트 픽스쳐 3.5mm", 120_000),
    "FX002": ("임플란트 픽스쳐 4.0mm", 135_000),
    "FX003": ("어버트먼트 표준형", 45_000),
    "FX004": ("힐링캡 소형", 8_000),
    "FX005": ("커버스크류", 5_000),
    "FX006": ("보철물 세라믹 크라운", 180_000),
    "FX007": ("임플란트 드릴 3.5", 250_000),
    "FX008": ("임플란트 드릴 4.0", 260_000),
    "FX009": ("임플란트 모터 핸드피스", 850_000),
    "FX010": ("소모품 키트 스탠다드", 35_000),
}


# 거래처별 출고 시나리오
# (거래처코드, 거래처명, 담당자, 담당자이메일_라벨, 출고리스트)
# 출고리스트: [(품목코드, 출고수량, 반품수량), ...]
# 이메일 라벨: SUB1/SUB2/SUB3 — F2 데모에서 .env의 SUB_ACCOUNTS로 치환

SCENARIOS = [
    # ── 이상 거래처 5개 ──
    ("CUST001", "강남B치과", "김수영", "SUB1", [
        ("FX001", 156, 0),   # 50 초과, 반품 0%
        ("FX002", 120, 0),   # 50 초과, 반품 0%
        ("FX003", 89, 0),    # 50 초과, 반품 0%
        ("FX004", 45, 2),    # 정상 범위
    ]),
    ("CUST002", "판교Y치과", "이정호", "SUB2", [
        ("FX005", 110, 0),   # 50 초과, 반품 0%
        ("FX006", 78, 0),    # 50 초과, 반품 0%
        ("FX010", 30, 1),
    ]),
    ("CUST003", "잠실Z치과", "박민지", "SUB3", [
        ("FX002", 95, 22),   # 50 초과, 반품 23% (과다)
        ("FX003", 60, 9),    # 50 초과, 반품 15% (과다)
        ("FX004", 35, 2),
    ]),
    ("CUST004", "성수K치과", "정우성", "SUB1", [
        ("FX001", 60, 3),    # 50 초과, 반품 5% (정상)
        ("FX007", 55, 2),    # 50 초과, 반품 3.6% (정상)
        ("FX010", 25, 2),
    ]),
    ("CUST005", "홍대M치과", "최윤서", "SUB2", [
        ("FX003", 80, 0),    # 50 초과, 반품 0%
        ("FX004", 25, 1),
    ]),
    # ── 정상 거래처 15개 ──
    ("CUST006", "분당S치과", "한지혜", "SUB3", [
        ("FX001", 32, 3), ("FX003", 28, 2), ("FX004", 15, 1),
    ]),
    ("CUST007", "용산P치과", "송재민", "SUB1", [
        ("FX002", 24, 2), ("FX005", 18, 1), ("FX010", 12, 0),
    ]),
    ("CUST008", "서초A치과", "조은서", "SUB2", [
        ("FX001", 45, 4), ("FX006", 15, 1), ("FX009", 2, 0),
    ]),
    ("CUST009", "마포N치과", "윤태호", "SUB3", [
        ("FX003", 38, 3), ("FX004", 22, 2), ("FX007", 8, 1),
    ]),
    ("CUST010", "송파D치과", "배현주", "SUB1", [
        ("FX001", 26, 2), ("FX002", 20, 2),
    ]),
    ("CUST011", "동탄L치과", "임지훈", "SUB2", [
        ("FX005", 30, 3), ("FX010", 16, 1),
    ]),
    ("CUST012", "인천O치과", "강수빈", "SUB3", [
        ("FX002", 18, 2), ("FX004", 12, 1), ("FX008", 6, 0),
    ]),
    ("CUST013", "수원R치과", "오세린", "SUB1", [
        ("FX001", 40, 4), ("FX003", 25, 2), ("FX010", 10, 1),
    ]),
    ("CUST014", "일산C치과", "홍대영", "SUB2", [
        ("FX006", 8, 1), ("FX003", 35, 3),
    ]),
    ("CUST015", "청담H치과", "서경민", "SUB3", [
        ("FX001", 20, 2), ("FX002", 22, 2), ("FX009", 1, 0),
    ]),
    ("CUST016", "광명T치과", "유나라", "SUB1", [
        ("FX005", 28, 3), ("FX010", 14, 2),
    ]),
    ("CUST017", "노원E치과", "남궁훈", "SUB2", [
        ("FX003", 32, 4), ("FX007", 6, 0),
    ]),
    ("CUST018", "구로G치과", "변지원", "SUB3", [
        ("FX001", 36, 3), ("FX004", 18, 1),
    ]),
    ("CUST019", "강동J치과", "전수경", "SUB1", [
        ("FX002", 42, 5), ("FX006", 5, 0),
    ]),
    ("CUST020", "양재F치과", "차영미", "SUB2", [
        ("FX001", 30, 2), ("FX005", 22, 2), ("FX010", 8, 1),
    ]),
]


def expand_scenarios() -> list[dict]:
    """시나리오를 개별 거래 행으로 펼침 (날짜 분산)"""
    random.seed(20260319)
    rows: list[dict] = []
    date_pool = [f"2026-03-{d:02d}" for d in range(1, 20)]  # 3/1 ~ 3/19

    for cust_code, cust_name, contact, email_label, shipments in SCENARIOS:
        for item_code, qty_total, return_total in shipments:
            item_name, unit_price = ITEMS[item_code]
            # 총 수량을 2~5개 거래로 분할
            n_splits = random.randint(2, 5) if qty_total > 20 else random.randint(1, 2)
            splits = split_qty(qty_total, n_splits)
            return_splits = split_qty(return_total, n_splits) if return_total > 0 else [0] * n_splits
            # 날짜 랜덤 배정
            dates = sorted(random.sample(date_pool, min(n_splits, len(date_pool))))
            for ship_date, qty, ret_qty in zip(dates, splits, return_splits):
                rows.append({
                    "거래일": ship_date,
                    "거래처코드": cust_code,
                    "거래처명": cust_name,
                    "품목코드": item_code,
                    "품목명": item_name,
                    "출고수량": qty,
                    "반품수량": ret_qty,
                    "단가": unit_price,
                    "금액": qty * unit_price,
                    "담당자": contact,
                    "담당자이메일": f"{email_label}@demo.local",
                })
    rows.sort(key=lambda r: (r["거래일"], r["거래처코드"], r["품목코드"]))
    return rows


def split_qty(total: int, n: int) -> list[int]:
    """정수를 n등분(대략)"""
    if n <= 1 or total == 0:
        return [total] + [0] * (n - 1) if n > 1 else [total]
    base = total // n
    rem = total - base * n
    result = [base] * n
    for i in range(rem):
        result[i] += 1
    random.shuffle(result)
    return result


def save_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "출고실적"

    headers = ["거래일", "거래처코드", "거래처명", "품목코드", "품목명",
               "출고수량", "반품수량", "단가", "금액", "담당자", "담당자이메일"]
    ws.append(headers)

    # 헤더 서식
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r[h] for h in headers])

    # 컬럼 너비
    widths = [12, 12, 14, 10, 24, 10, 10, 10, 14, 10, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 숫자 서식 (천단위 콤마)
    for row_idx in range(2, len(rows) + 2):
        for col_name in ["출고수량", "반품수량", "단가", "금액"]:
            col_idx = headers.index(col_name) + 1
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    # 틀 고정
    ws.freeze_panes = "A2"

    wb.save(OUT_PATH)


def print_summary(rows: list[dict]) -> None:
    # 거래처별 요약
    from collections import defaultdict
    by_cust: dict[str, dict] = defaultdict(lambda: {"name": "", "items": defaultdict(lambda: [0, 0])})
    for r in rows:
        cust = r["거래처코드"]
        by_cust[cust]["name"] = r["거래처명"]
        it = by_cust[cust]["items"][r["품목코드"]]
        it[0] += r["출고수량"]
        it[1] += r["반품수량"]

    print(f"[시작] 이상거래 Mock 데이터 생성 → {OUT_PATH.name}")
    print()
    print(f"  총 거래처: {len(by_cust)}개")
    print(f"  총 거래 행: {len(rows)}행")
    print()
    print(f"  [이상 의심 거래처 - 품목별 50개 초과 출고]")
    for cust in sorted(by_cust.keys()):
        info = by_cust[cust]
        over_50 = [(code, q, rq) for code, (q, rq) in info["items"].items() if q > 50]
        if over_50:
            total_ret = sum(rq for _, _, rq in over_50)
            total_out = sum(q for _, q, _ in over_50)
            ret_rate = (total_ret / total_out * 100) if total_out else 0
            mark = ""
            if ret_rate == 0 and len(over_50) >= 2:
                mark = "  <= 이상품목 다수 + 반품 0% [높음]"
            elif ret_rate == 0:
                mark = "  <= 반품률 0% [주의]"
            elif ret_rate > 10:
                mark = f"  <= 반품률 {ret_rate:.0f}% [주의]"
            print(f"    {cust} {info['name']:<10} - 이상품목 {len(over_50)}개, 반품률 {ret_rate:>4.1f}%{mark}")


def main() -> None:
    rows = expand_scenarios()
    save_xlsx(rows)
    print_summary(rows)
    print()
    print(f"[완료] {OUT_PATH}")


if __name__ == "__main__":
    main()
