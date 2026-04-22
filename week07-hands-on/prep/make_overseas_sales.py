"""
해외법인 40개 Mock 데이터 생성 (D 데모: PPT+HTML+JSX 3종 대시보드)

육아름 차장님 업무 맥락:
- 40개 해외법인 월별 매출/원가/판관비 취합
- 법인 통화 → KRW 환산
- 지역별·품목군별·법인별 다차원 분석

출력 xlsx 구성:
- Sheet 1 (법인마스터): 40법인 고정정보 (법인코드·국가·지역·통화·담당자)
- Sheet 2 (월별매출): 40 × 6개월 × 3품목군 = 720행
- Sheet 3 (환율): 26개 통화 월말 KRW 환율

실행:
  py week07-hands-on/prep/make_overseas_sales.py

출력:
  week07-hands-on/demos/dashboard_3format/data/overseas_sales.xlsx
"""

from __future__ import annotations

import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


HERE = Path(__file__).parent
OUT_DIR = HERE.parent / "demos" / "dashboard_3format" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / "overseas_sales.xlsx"


# (법인코드, 법인명, 국가, 지역, 통화, 담당자, 월매출_기준_USD환산)
LEGIONS = [
    # Americas (10)
    ("US01", "미국 동부법인",      "미국",     "Americas", "USD", "Jennifer Park",   1_800_000),
    ("US02", "미국 서부법인",      "미국",     "Americas", "USD", "Daniel Kim",      1_500_000),
    ("US03", "미국 중부법인",      "미국",     "Americas", "USD", "Michael Lee",     1_100_000),
    ("US04", "미국 남부법인",      "미국",     "Americas", "USD",   "Sophia Chen",     780_000),
    ("US05", "미국 R&D센터",       "미국",     "Americas", "USD", "David Wilson",      950_000),
    ("CA01", "캐나다 토론토법인",  "캐나다",   "Americas", "CAD", "Emily Brown",       520_000),
    ("MX01", "멕시코시티법인",     "멕시코",   "Americas", "MXN", "Carlos Ruiz",       380_000),
    ("BR01", "브라질 상파울로",    "브라질",   "Americas", "BRL", "Rafael Silva",      610_000),
    ("BR02", "브라질 리우데자네이루", "브라질", "Americas", "BRL", "Lucas Oliveira",    420_000),
    ("CL01", "칠레 산티아고",      "칠레",     "Americas", "CLP", "Isabella Torres",   260_000),

    # APAC (15)
    ("CN01", "상해법인",           "중국",     "APAC",     "CNY", "王丽",              1_700_000),
    ("CN02", "북경법인",           "중국",     "APAC",     "CNY", "张伟",              1_400_000),
    ("CN03", "광주법인",           "중국",     "APAC",     "CNY", "李娜",                920_000),
    ("JP01", "도쿄법인",           "일본",     "APAC",     "JPY", "田中 健太",         1_250_000),
    ("JP02", "오사카법인",         "일본",     "APAC",     "JPY", "佐藤 美咲",           680_000),
    ("HK01", "홍콩법인",           "홍콩",     "APAC",     "HKD", "Wong Ka Ho",        890_000),
    ("SG01", "싱가포르법인",       "싱가포르", "APAC",     "SGD", "Tan Wei Ming",      780_000),
    ("TW01", "대만 타이페이법인",  "대만",     "APAC",     "TWD", "陳俊宏",              540_000),
    ("VN01", "베트남 하노이법인",  "베트남",   "APAC",     "VND", "Nguyen Van An",     340_000),
    ("VN02", "베트남 호치민법인",  "베트남",   "APAC",     "VND", "Tran Thi Lan",      420_000),
    ("TH01", "태국 방콕법인",      "태국",     "APAC",     "THB", "Somchai Wong",      380_000),
    ("ID01", "인도네시아 자카르타","인도네시아","APAC",    "IDR", "Budi Santoso",      290_000),
    ("MY01", "말레이시아 쿠알라룸푸르","말레이시아","APAC","MYR", "Ahmad Hafiz",       310_000),
    ("IN01", "인도 뭄바이법인",    "인도",     "APAC",     "INR", "Priya Sharma",      510_000),
    ("IN02", "인도 뉴델리법인",    "인도",     "APAC",     "INR", "Rajesh Kumar",      380_000),

    # EMEA (12)
    ("DE01", "독일 프랑크푸르트",  "독일",     "EMEA",     "EUR", "Hans Mueller",      1_350_000),
    ("DE02", "독일 뮌헨법인",      "독일",     "EMEA",     "EUR", "Anna Schmidt",        870_000),
    ("GB01", "영국 런던법인",      "영국",     "EMEA",     "GBP", "James Smith",       1_100_000),
    ("FR01", "프랑스 파리법인",    "프랑스",   "EMEA",     "EUR", "Sophie Martin",       780_000),
    ("NL01", "네덜란드 암스테르담","네덜란드", "EMEA",     "EUR", "Jan de Vries",        590_000),
    ("IT01", "이탈리아 밀라노",    "이탈리아", "EMEA",     "EUR", "Marco Rossi",         540_000),
    ("ES01", "스페인 마드리드",    "스페인",   "EMEA",     "EUR", "Elena Garcia",        460_000),
    ("PL01", "폴란드 바르샤바",    "폴란드",   "EMEA",     "PLN", "Jakub Nowak",         320_000),
    ("TR01", "튀르키예 이스탄불",  "튀르키예", "EMEA",     "TRY", "Mehmet Yilmaz",       270_000),
    ("AE01", "UAE 두바이법인",     "UAE",      "EMEA",     "AED", "Ahmed Al-Rashid",     640_000),
    ("RU01", "러시아 모스크바",    "러시아",   "EMEA",     "RUB", "Aleksandr Ivanov",    380_000),
    ("ZA01", "남아공 요하네스버그","남아공",   "EMEA",     "ZAR", "Thabo Nkosi",         240_000),

    # Oceania (3)
    ("AU01", "호주 시드니법인",    "호주",     "Oceania",  "AUD", "Oliver Taylor",       720_000),
    ("AU02", "호주 멜버른법인",    "호주",     "Oceania",  "AUD", "Charlotte Wilson",    480_000),
    ("NZ01", "뉴질랜드 오클랜드",  "뉴질랜드", "Oceania",  "NZD", "William Anderson",    310_000),
]


# 월말 KRW 환율 (가상 — week05 패턴 기반 임의 설정)
EXCHANGE_RATES = {
    "USD": 1_350.0,
    "JPY": 9.2,
    "CNY": 186.0,
    "EUR": 1_480.0,
    "VND": 0.056,
    "INR": 16.0,
    "GBP": 1_720.0,
    "THB": 38.9,
    "HKD": 172.0,
    "SGD": 1_000.0,
    "TWD": 43.0,
    "IDR": 0.085,
    "MYR": 290.0,
    "CAD": 990.0,
    "MXN": 68.0,
    "BRL": 240.0,
    "CLP": 1.5,
    "AUD": 890.0,
    "NZD": 820.0,
    "PLN": 335.0,
    "TRY": 35.0,
    "AED": 367.0,
    "RUB": 14.0,
    "ZAR": 73.0,
    # 참고: PHP는 법인 제거로 미사용
}

MONTHS = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]

# 품목군 (비중)
PRODUCT_GROUPS = [
    ("제품A 임플란트 본체", 0.60),
    ("제품B 소모품",        0.30),
    ("서비스 교육·컨설팅",  0.10),
]


def month_end_date(month: str) -> str:
    """YYYY-MM → YYYY-MM-DD (월말)"""
    year, m = map(int, month.split("-"))
    if m == 2:
        day = 28
    elif m in (4, 6, 9, 11):
        day = 30
    else:
        day = 31
    return f"{year}-{m:02d}-{day:02d}"


def build_legion_master_sheet(ws) -> None:
    headers = ["법인코드", "법인명", "국가", "지역", "통화", "담당자", "이메일"]
    ws.append(headers)
    for code, name, country, region, currency, contact, _ in LEGIONS:
        # 이메일은 법인코드 기반 placeholder (실명 아님)
        email = f"{code.lower()}@demo.local"
        ws.append([code, name, country, region, currency, contact, email])
    style_header(ws, len(headers))
    widths = [10, 28, 12, 10, 8, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def build_monthly_sales_sheet(ws) -> None:
    headers = ["법인코드", "법인명", "연월", "품목군", "매출(현지통화)", "통화",
               "원가율", "원가", "판관비율", "판관비", "영업이익"]
    ws.append(headers)

    random.seed(20260319)
    rows = []

    for code, name, country, region, currency, contact, base_usd in LEGIONS:
        for month in MONTHS:
            # 월별 노이즈 ±15%
            month_factor = random.uniform(0.85, 1.15)
            # USD → 현지통화: 현지 = USD × (USD_to_KRW / 현지_to_KRW) = USD × (1350 / rate[currency])
            total_local = base_usd * month_factor * (1_350.0 / EXCHANGE_RATES[currency])
            # 단, 통화별로 보기좋은 반올림
            if currency in ("JPY", "VND", "IDR", "CLP", "KRW"):
                total_local = round(total_local / 1000) * 1000  # 1천 단위 반올림
            elif currency in ("INR", "THB", "PHP", "RUB", "TWD", "MXN", "TRY", "ZAR"):
                total_local = round(total_local / 100) * 100
            else:
                total_local = round(total_local)

            for pg_name, pg_weight in PRODUCT_GROUPS:
                # 품목군별 비중에 ±5%p 노이즈
                pg_noise = random.uniform(-0.05, 0.05)
                actual_weight = max(0.01, pg_weight + pg_noise)
                pg_sales = total_local * actual_weight

                # 원가율 60 ± 8%
                cost_rate = round(random.uniform(0.52, 0.68), 3)
                cost = pg_sales * cost_rate

                # 판관비율 12 ± 4%
                sga_rate = round(random.uniform(0.08, 0.16), 3)
                sga = pg_sales * sga_rate

                op_profit = pg_sales - cost - sga

                # 숫자 반올림
                if currency in ("JPY", "VND", "IDR", "CLP"):
                    pg_sales = round(pg_sales)
                    cost = round(cost)
                    sga = round(sga)
                    op_profit = round(op_profit)
                else:
                    pg_sales = round(pg_sales, 2)
                    cost = round(cost, 2)
                    sga = round(sga, 2)
                    op_profit = round(op_profit, 2)

                rows.append([code, name, month, pg_name, pg_sales, currency,
                             cost_rate, cost, sga_rate, sga, op_profit])

    for row in rows:
        ws.append(row)

    style_header(ws, len(headers))
    widths = [10, 28, 10, 26, 18, 8, 10, 16, 10, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 숫자 서식
    for row_idx in range(2, len(rows) + 2):
        for col_name in ["매출(현지통화)", "원가", "판관비", "영업이익"]:
            col_idx = headers.index(col_name) + 1
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"
        for col_name in ["원가율", "판관비율"]:
            col_idx = headers.index(col_name) + 1
            ws.cell(row=row_idx, column=col_idx).number_format = "0.0%"

    ws.freeze_panes = "A2"
    return len(rows)


def build_exchange_rates_sheet(ws) -> None:
    headers = ["통화", "환율(KRW)", "기준일"]
    ws.append(headers)
    # 각 월말 6개 * 26통화 = 156행 (여기서는 최신 1시점만 간단히)
    # 실제 시연에서는 최신 환율만 쓰니까 6월말 기준 1건씩만 저장
    for currency, rate in sorted(EXCHANGE_RATES.items()):
        ws.append([currency, rate, "2026-06-30"])

    style_header(ws, len(headers))
    widths = [10, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row_idx in range(2, len(EXCHANGE_RATES) + 2):
        ws.cell(row=row_idx, column=2).number_format = "#,##0.000"

    ws.freeze_panes = "A2"


def style_header(ws, ncols: int) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for c_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def main() -> None:
    wb = Workbook()

    # Sheet 1: 법인마스터
    ws1 = wb.active
    ws1.title = "법인마스터"
    build_legion_master_sheet(ws1)

    # Sheet 2: 월별매출
    ws2 = wb.create_sheet("월별매출")
    row_count = build_monthly_sales_sheet(ws2)

    # Sheet 3: 환율
    ws3 = wb.create_sheet("환율")
    build_exchange_rates_sheet(ws3)

    wb.save(OUT_PATH)

    print(f"[시작] 해외법인 Mock 데이터 생성 → {OUT_PATH.name}")
    print()
    print(f"  Sheet 1 법인마스터 : {len(LEGIONS)}개 법인")
    print(f"    - Americas  : {sum(1 for l in LEGIONS if l[3] == 'Americas')}개")
    print(f"    - APAC      : {sum(1 for l in LEGIONS if l[3] == 'APAC')}개")
    print(f"    - EMEA      : {sum(1 for l in LEGIONS if l[3] == 'EMEA')}개")
    print(f"    - Oceania   : {sum(1 for l in LEGIONS if l[3] == 'Oceania')}개")
    print(f"  Sheet 2 월별매출   : {row_count}행 ({len(LEGIONS)}법인 x {len(MONTHS)}개월 x {len(PRODUCT_GROUPS)}품목군)")
    print(f"  Sheet 3 환율       : {len(EXCHANGE_RATES)}개 통화")
    print()
    print(f"[완료] {OUT_PATH}")


if __name__ == "__main__":
    main()
