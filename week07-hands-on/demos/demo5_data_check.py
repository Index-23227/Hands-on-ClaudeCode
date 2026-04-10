"""
데모 5: 데이터 정합성 체크 — 두 시스템 데이터 불일치 자동 탐지
"눈으로 대조하던 감사 자료, 3초에 끝냅니다"
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "output")
os.makedirs(OUT, exist_ok=True)


# ── 1. 샘플: 회계시스템 vs ERP 데이터 ──
def create_sample():
    # 회계시스템 데이터
    accounting = [
        ("2026-04-01", "미국법인", "매출", 141000),
        ("2026-04-05", "일본법인", "매출", 15500000),
        ("2026-04-08", "중국법인", "매출", 960000),
        ("2026-04-10", "독일법인", "매출", 112000),
        ("2026-04-12", "베트남법인", "매출", 3600000000),
        ("2026-04-15", "인도법인", "매출", 9500000),
        ("2026-04-18", "영국법인", "매출", 81000),
        ("2026-04-20", "태국법인", "매출", 4700000),
    ]

    # ERP 데이터 — 의도적으로 3건 불일치
    erp = [
        ("2026-04-01", "미국법인", "매출", 141000),      # OK
        ("2026-04-05", "일본법인", "매출", 15500000),     # OK
        ("2026-04-08", "중국법인", "매출", 950000),       # 불일치! (960000 vs 950000)
        ("2026-04-10", "독일법인", "매출", 112000),       # OK
        ("2026-04-12", "베트남법인", "매출", 3600000000), # OK
        ("2026-04-15", "인도법인", "매출", 9800000),      # 불일치! (9500000 vs 9800000)
        ("2026-04-18", "영국법인", "매출", 81000),        # OK
        # 태국법인 누락!                                    # 불일치! (누락)
    ]

    for name, data in [("회계시스템", accounting), ("ERP", erp)]:
        wb = Workbook()
        ws = wb.active
        ws.title = "매출"
        ws.append(["일자", "법인", "계정", "금액"])
        for row in data:
            ws.append(list(row))
        wb.save(os.path.join(OUT, f"{name}_4월.xlsx"))

    return accounting, erp


def check_integrity(accounting, erp):
    acc_dict = {(r[0], r[1]): r[3] for r in accounting}
    erp_dict = {(r[0], r[1]): r[3] for r in erp}

    issues = []

    # 회계에 있는데 ERP에 없는 것
    for key in acc_dict:
        if key not in erp_dict:
            issues.append({
                "유형": "ERP 누락",
                "일자": key[0],
                "법인": key[1],
                "회계": f"{acc_dict[key]:,.0f}",
                "ERP": "—",
                "차이": f"{acc_dict[key]:,.0f}",
            })
        elif acc_dict[key] != erp_dict[key]:
            diff = erp_dict[key] - acc_dict[key]
            issues.append({
                "유형": "금액 불일치",
                "일자": key[0],
                "법인": key[1],
                "회계": f"{acc_dict[key]:,.0f}",
                "ERP": f"{erp_dict[key]:,.0f}",
                "차이": f"{diff:+,.0f}",
            })

    # ERP에 있는데 회계에 없는 것
    for key in erp_dict:
        if key not in acc_dict:
            issues.append({
                "유형": "회계 누락",
                "일자": key[0],
                "법인": key[1],
                "회계": "—",
                "ERP": f"{erp_dict[key]:,.0f}",
                "차이": f"{erp_dict[key]:,.0f}",
            })

    return issues


def save_report(issues, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "정합성 체크"

    headers = ["유형", "일자", "법인", "회계시스템", "ERP", "차이"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C0392B")

    red_fill = PatternFill("solid", fgColor="FADBD8")
    for issue in issues:
        ws.append([issue["유형"], issue["일자"], issue["법인"],
                   issue["회계"], issue["ERP"], issue["차이"]])
        for cell in ws[ws.max_row]:
            cell.fill = red_fill

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 16

    wb.save(out_path)


if __name__ == "__main__":
    print("=== 데모 5: 데이터 정합성 체크 ===")
    print()

    accounting, erp = create_sample()
    print(f"  회계시스템: {len(accounting)}건")
    print(f"  ERP:       {len(erp)}건")

    issues = check_integrity(accounting, erp)
    print(f"\n  불일치 발견: {len(issues)}건")
    print(f"  {'─'*60}")
    for issue in issues:
        print(f"  [{issue['유형']:8s}] {issue['일자']} {issue['법인']:8s} | 회계: {issue['회계']:>14s} | ERP: {issue['ERP']:>14s} | 차이: {issue['차이']}")

    out = os.path.join(OUT, "정합성_체크결과.xlsx")
    save_report(issues, out)
    print(f"\n  결과 저장: {out}")
    print("  → 불일치 행이 빨간색으로 표시됩니다.")
    print("  → 감사 자료 대조를 눈이 아니라 코드가 합니다.")
