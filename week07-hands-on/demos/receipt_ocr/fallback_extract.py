"""
데모 A 폴백 — Claude Vision이 영수증 8장을 읽은 결과를 엑셀로 정리

실제 시연에서 Claude가 이미지를 처리해야 하지만, 멀티모달 기능이 실패하거나
시간이 부족할 때 폴백으로 실행. 수강생이 보게 될 최종 엑셀과 동일 형식.

실행:
  py week07-hands-on/demos/receipt_ocr/fallback_extract.py

출력:
  week07-hands-on/demos/receipt_ocr/output/경비정리_2026-03.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


HERE = Path(__file__).parent
OUT_DIR = HERE / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "경비정리_2026-03.xlsx"


# 영수증 8장에서 Claude Vision이 추출할 예상 결과
# (실제 시연에서는 Claude가 이미지를 읽어서 이 표를 만들어냄)
RESULTS = [
    ("2026-03-05", "스타벅스 강남역점",     "아메리카노 등 음료",   18_200,  "법인카드"),
    ("2026-03-08", "GS25 역삼중앙점",       "간식·음료",            12_900,  "개인카드"),
    ("2026-03-12", "GS칼텍스 양재주유소",   "휘발유 45.2L",         68_900,  "법인카드"),
    ("2026-03-15", "오피스디포 여의도점",   "A4용지·토너 등",       189_700, "법인카드"),
    ("2026-03-18", "본죽 여의도점",         "전복죽 2인 외",        40_000,  "개인카드"),
    ("2026-03-19", "개인택시",              "서울역→여의도",        8_700,   "법인카드"),
    ("2026-03-20", "코레일 KTX",            "서울→부산 특실",       79_800,  "법인카드"),
    ("2026-03-21", "베스트웨스턴 해운대",    "숙박 1박 + 조식",      196_000, "법인카드"),
]

HEADERS = ["날짜", "매장명", "항목 요약", "합계금액", "결제방식"]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
TOTAL_FONT = Font(bold=True)


def style_header_row(ws, row_idx: int) -> None:
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")


def style_total_row(ws, row_idx: int) -> None:
    for col_idx in range(1, len(HEADERS) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT


def apply_column_widths(ws) -> None:
    widths = [12, 26, 22, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def build_sheet1_all(ws) -> None:
    ws.title = "전체 목록"
    ws.append(HEADERS)
    style_header_row(ws, 1)

    # 날짜순 정렬
    sorted_results = sorted(RESULTS, key=lambda r: r[0])
    for row in sorted_results:
        ws.append(list(row))

    # 총계 행
    total = sum(r[3] for r in sorted_results)
    ws.append(["", "", "총계", total, ""])
    style_total_row(ws, ws.max_row)

    # 숫자 서식
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = "#,##0"

    apply_column_widths(ws)
    ws.freeze_panes = "A2"


def build_sheet2_grouped(ws) -> None:
    ws.title = "결제방식별"
    ws.append(HEADERS)
    style_header_row(ws, 1)

    # 법인카드 섹션
    corp_rows = [r for r in RESULTS if r[4] == "법인카드"]
    for row in sorted(corp_rows, key=lambda r: r[0]):
        ws.append(list(row))
    corp_total = sum(r[3] for r in corp_rows)
    ws.append(["", "", "법인카드 소계", corp_total, f"{len(corp_rows)}건"])
    style_total_row(ws, ws.max_row)

    # 빈 행 구분
    ws.append([""] * len(HEADERS))

    # 개인카드 섹션
    personal_rows = [r for r in RESULTS if r[4] == "개인카드"]
    for row in sorted(personal_rows, key=lambda r: r[0]):
        ws.append(list(row))
    personal_total = sum(r[3] for r in personal_rows)
    ws.append(["", "", "개인카드 소계", personal_total, f"{len(personal_rows)}건"])
    style_total_row(ws, ws.max_row)

    # 빈 행
    ws.append([""] * len(HEADERS))

    # 전체 합계
    ws.append(["", "", "전체 합계", corp_total + personal_total, f"{len(RESULTS)}건"])
    style_total_row(ws, ws.max_row)

    # 숫자 서식
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=4)
        if cell.value is not None and cell.value != "":
            cell.number_format = "#,##0"

    apply_column_widths(ws)
    ws.freeze_panes = "A2"


def main() -> None:
    wb = Workbook()
    ws1 = wb.active
    build_sheet1_all(ws1)

    ws2 = wb.create_sheet("결제방식별")
    build_sheet2_grouped(ws2)

    wb.save(OUT_PATH)

    corp_total = sum(r[3] for r in RESULTS if r[4] == "법인카드")
    personal_total = sum(r[3] for r in RESULTS if r[4] == "개인카드")

    print(f"[완료] {OUT_PATH}")
    print()
    print(f"  시트1 전체 목록 : {len(RESULTS)}건")
    print(f"  시트2 결제방식별:")
    print(f"    - 법인카드 소계 : {sum(1 for r in RESULTS if r[4] == '법인카드'):>2}건  {corp_total:>10,}원")
    print(f"    - 개인카드 소계 : {sum(1 for r in RESULTS if r[4] == '개인카드'):>2}건  {personal_total:>10,}원")
    print(f"    - 전체 합계     : {len(RESULTS):>2}건  {corp_total + personal_total:>10,}원")


if __name__ == "__main__":
    main()
