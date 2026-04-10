"""
데모 2: PDF에서 숫자 뽑기 — 세금계산서 PDF → 엑셀 정리
"손으로 옮기던 걸 자동으로"

※ 이 데모는 PDF 파일 대신 텍스트 시뮬레이션으로 보여줍니다.
  실제 업무에서는 Claude Code에게 "이 PDF 읽어서 엑셀로 정리해줘"라고 시키면
  pdfplumber/PyPDF2 등을 사용해 동일하게 처리합니다.
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "output")
os.makedirs(OUT, exist_ok=True)

# ── 1. PDF에서 추출한 텍스트 시뮬레이션 ──
# 실제로는 pdfplumber.open("세금계산서.pdf").pages[0].extract_text() 결과
SAMPLE_INVOICES = [
    """
전자세금계산서
공급자: (주)한국전자   사업자번호: 123-45-67890
공급받는자: 글로벌테크(주)   사업자번호: 987-65-43210
작성일자: 2026-04-15
품목: 반도체 부품 A-100
공급가액: 45,000,000원
세액: 4,500,000원
합계금액: 49,500,000원
""",
    """
전자세금계산서
공급자: 대한물산(주)   사업자번호: 111-22-33344
공급받는자: 글로벌테크(주)   사업자번호: 987-65-43210
작성일자: 2026-04-18
품목: 포장재 B-200
공급가액: 12,500,000원
세액: 1,250,000원
합계금액: 13,750,000원
""",
    """
전자세금계산서
공급자: 서울운송(주)   사업자번호: 555-66-77788
공급받는자: 글로벌테크(주)   사업자번호: 987-65-43210
작성일자: 2026-04-22
품목: 국내 운송료
공급가액: 3,200,000원
세액: 320,000원
합계금액: 3,520,000원
""",
]


def extract_field(text, pattern):
    m = re.search(pattern, text)
    return m.group(1).strip() if m else ""


def parse_amount(s):
    return int(re.sub(r"[,원\s]", "", s)) if s else 0


def extract_invoice(text):
    return {
        "공급자": extract_field(text, r"공급자:\s*(.+?)(?:\s{2,}|사업자)"),
        "사업자번호": extract_field(text, r"공급자.*?사업자번호:\s*([\d-]+)"),
        "작성일자": extract_field(text, r"작성일자:\s*(\d{4}-\d{2}-\d{2})"),
        "품목": extract_field(text, r"품목:\s*(.+)"),
        "공급가액": parse_amount(extract_field(text, r"공급가액:\s*([\d,]+)")),
        "세액": parse_amount(extract_field(text, r"세액:\s*([\d,]+)")),
        "합계": parse_amount(extract_field(text, r"합계금액:\s*([\d,]+)")),
    }


def save_to_excel(invoices, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "세금계산서 추출"

    headers = ["No", "작성일자", "공급자", "사업자번호", "품목", "공급가액", "세액", "합계금액"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")
        cell.alignment = Alignment(horizontal="center")

    for i, inv in enumerate(invoices, 1):
        ws.append([
            i, inv["작성일자"], inv["공급자"], inv["사업자번호"],
            inv["품목"], inv["공급가액"], inv["세액"], inv["합계"],
        ])

    # 합계 행
    last = len(invoices) + 2
    ws.append(["", "", "", "", "합계",
               f"=SUM(F2:F{last-1})", f"=SUM(G2:G{last-1})", f"=SUM(H2:H{last-1})"])
    for cell in ws[last]:
        cell.font = Font(bold=True)

    # 숫자 서식
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = '#,##0'

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 18

    wb.save(out_path)


if __name__ == "__main__":
    print("=== 데모 2: PDF → 엑셀 추출 ===")
    print()
    print("  [시뮬레이션] 세금계산서 3장에서 데이터 추출 중...")
    print()

    invoices = []
    for i, text in enumerate(SAMPLE_INVOICES, 1):
        inv = extract_invoice(text)
        invoices.append(inv)
        print(f"  #{i} {inv['작성일자']} | {inv['공급자']:12s} | {inv['공급가액']:>12,}원 + 세액 {inv['세액']:>10,}원 = {inv['합계']:>12,}원")

    total = sum(inv["합계"] for inv in invoices)
    print(f"  {'─'*70}")
    print(f"  합계: {total:>58,}원")

    out_path = os.path.join(OUT, "세금계산서_추출.xlsx")
    save_to_excel(invoices, out_path)
    print(f"\n  결과 저장: {out_path}")
    print("  → 실제 업무에서는 Claude에게 'PDF 파일 읽어서 이 형식으로 정리해줘'라고 시키면 됩니다.")
