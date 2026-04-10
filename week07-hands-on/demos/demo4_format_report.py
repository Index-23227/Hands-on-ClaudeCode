"""
데모 4: 엑셀 서식 자동화 — 날것 데이터 → 보고서 형태 엑셀
"매번 30분 걸리던 포맷팅을 자동으로"
"""
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "output")
DB_PATH = os.path.normpath(os.path.join(HERE, "..", "..", "week05-hands-on", "data", "sales.db"))
os.makedirs(OUT, exist_ok=True)

# 스타일 정의
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="ECF0F1")
TOTAL_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="2C3E50")
THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)


def generate_report(db_path, out_path):
    conn = sqlite3.connect(db_path)

    # 데이터 조회
    corps = conn.execute("""
        SELECT c.corp_name, c.currency, SUM(m.amount) AS total,
               CAST(ROUND(SUM(m.amount) * r.rate) AS INT) AS krw
        FROM corporations c
        JOIN monthly_sales m ON m.corp_code = c.corp_code
        JOIN exchange_rates r ON r.currency = c.currency
         AND r.rate_date = (SELECT MAX(rate_date) FROM exchange_rates WHERE currency = c.currency)
        GROUP BY c.corp_code
        ORDER BY krw DESC
    """).fetchall()

    months = conn.execute("""
        SELECT DISTINCT month FROM monthly_sales ORDER BY month
    """).fetchall()
    month_list = [m[0] for m in months]

    monthly = conn.execute("""
        SELECT c.corp_name, m.month, m.amount
        FROM corporations c
        JOIN monthly_sales m ON m.corp_code = c.corp_code
        ORDER BY c.corp_name, m.month
    """).fetchall()
    conn.close()

    wb = Workbook()

    # ── 시트 1: 법인별 요약 ──
    ws1 = wb.active
    ws1.title = "법인별 요약"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "2026년 상반기 법인별 매출 보고서"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = ""  # 빈 줄

    headers = ["법인명", "통화", "외화 합계", "원화 환산"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, (name, currency, total, krw) in enumerate(corps):
        row = 4 + i
        ws1.cell(row, 1, name)
        ws1.cell(row, 2, currency).alignment = Alignment(horizontal="center")
        ws1.cell(row, 3, total).number_format = '#,##0'
        ws1.cell(row, 4, krw).number_format = '#,##0'

    # 합계 행
    total_row = 4 + len(corps)
    ws1.cell(total_row, 1, "합계").font = TOTAL_FONT
    ws1.cell(total_row, 1).fill = TOTAL_FILL
    for col in range(2, 5):
        ws1.cell(total_row, col).fill = TOTAL_FILL
    ws1.cell(total_row, 4, sum(c[3] for c in corps)).number_format = '#,##0'
    ws1.cell(total_row, 4).font = TOTAL_FONT

    # 테두리
    for row in ws1.iter_rows(min_row=3, max_row=total_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = THIN_BORDER

    # 열 너비
    for col, w in zip("ABCD", [14, 8, 18, 18]):
        ws1.column_dimensions[col].width = w

    # 차트
    chart = BarChart()
    chart.title = "법인별 원화 환산 매출"
    chart.y_axis.title = "원"
    chart.x_axis.title = "법인"
    chart.style = 10
    data = Reference(ws1, min_col=4, min_row=3, max_row=total_row - 1)
    cats = Reference(ws1, min_col=1, min_row=4, max_row=total_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 20
    chart.height = 12
    ws1.add_chart(chart, f"A{total_row + 2}")

    # ── 시트 2: 월별 상세 ──
    ws2 = wb.create_sheet("월별 상세")
    ws2.merge_cells(f"A1:{get_column_letter(2+len(month_list))}1")
    ws2["A1"] = "월별 매출 상세"
    ws2["A1"].font = TITLE_FONT

    headers2 = ["법인명"] + month_list + ["합계"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # 월별 데이터 pivot
    pivot = {}
    for name, month, amount in monthly:
        pivot.setdefault(name, {})[month] = amount

    for i, name in enumerate(sorted(pivot.keys())):
        row = 4 + i
        ws2.cell(row, 1, name)
        row_total = 0
        for j, m in enumerate(month_list):
            val = pivot[name].get(m, 0)
            ws2.cell(row, 2 + j, val).number_format = '#,##0'
            row_total += val
        ws2.cell(row, 2 + len(month_list), row_total).number_format = '#,##0'
        ws2.cell(row, 2 + len(month_list)).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 14
    for j in range(len(month_list) + 1):
        ws2.column_dimensions[get_column_letter(2 + j)].width = 16

    wb.save(out_path)


if __name__ == "__main__":
    print("=== 데모 4: 엑셀 서식 자동화 ===")

    if not os.path.exists(DB_PATH):
        print(f"  [ERROR] DB not found: {DB_PATH}")
        exit(1)

    out = os.path.join(OUT, "매출보고서_2026상반기.xlsx")
    generate_report(DB_PATH, out)
    print(f"  보고서 생성: {out}")
    print("  → 시트 1: 법인별 요약 + 차트")
    print("  → 시트 2: 월별 상세 피벗")
    print("  → 서식, 색상, 테두리, 합계행까지 자동 완성!")
    print()
    print("  여러분의 보고서 양식을 Claude에게 보여주고")
    print("  '이 형식으로 만들어줘'라고 시키면 됩니다.")
