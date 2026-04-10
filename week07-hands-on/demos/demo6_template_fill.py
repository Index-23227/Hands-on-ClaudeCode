"""
데모 6: 반복 보고서 템플릿 자동 채우기
"매월 같은 양식에 숫자만 바꾸는 일, 자동으로"
"""
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "output")
DB_PATH = os.path.normpath(os.path.join(HERE, "..", "..", "week05-hands-on", "data", "sales.db"))
os.makedirs(OUT, exist_ok=True)

# 스타일
THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
BLUE = PatternFill("solid", fgColor="D6EAF8")
GRAY = PatternFill("solid", fgColor="F2F3F4")
DARK = PatternFill("solid", fgColor="2C3E50")
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)


def fill_monthly_report(db_path, target_month, out_path):
    """target_month: '2026-04' 같은 형식"""
    conn = sqlite3.connect(db_path)

    # 해당 월 데이터
    rows = conn.execute("""
        SELECT c.corp_code, c.corp_name, c.currency, m.amount, m.note,
               r.rate,
               CAST(ROUND(m.amount * r.rate) AS INT) AS krw
        FROM corporations c
        LEFT JOIN monthly_sales m ON m.corp_code = c.corp_code AND m.month = ?
        LEFT JOIN exchange_rates r ON r.currency = c.currency
         AND r.rate_date = (SELECT MAX(rate_date) FROM exchange_rates WHERE currency = c.currency)
        ORDER BY c.corp_name
    """, (target_month,)).fetchall()

    # 전월 데이터 (비교용)
    year, mon = target_month.split("-")
    prev_month = f"{year}-{int(mon)-1:02d}" if int(mon) > 1 else f"{int(year)-1}-12"
    prev_rows = conn.execute("""
        SELECT c.corp_code, m.amount
        FROM corporations c
        LEFT JOIN monthly_sales m ON m.corp_code = c.corp_code AND m.month = ?
    """, (prev_month,)).fetchall()
    prev_dict = {r[0]: r[1] for r in prev_rows}

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"{target_month} 매출 보고서"

    # ── 타이틀 영역 ──
    ws.merge_cells("A1:G1")
    ws["A1"] = f"월별 매출 보고서 — {target_month}"
    ws["A1"].font = Font(bold=True, size=16, color="2C3E50")

    ws.merge_cells("A2:G2")
    ws["A2"] = f"작성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  자동 생성"
    ws["A2"].font = Font(size=9, color="7F8C8D")

    # ── 헤더 ──
    headers = ["법인명", "통화", "당월 매출", "전월 매출", "증감", "증감률", "원화 환산"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = WHITE_BOLD
        cell.fill = DARK
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    # ── 데이터 ──
    total_krw = 0
    for i, (code, name, currency, amount, note, rate, krw) in enumerate(rows):
        row_num = 5 + i
        amount = amount or 0
        krw = krw or 0
        prev = prev_dict.get(code) or 0
        diff = amount - prev
        pct = (diff / prev * 100) if prev else 0

        ws.cell(row_num, 1, name).border = THIN
        ws.cell(row_num, 2, currency).border = THIN
        ws.cell(row_num, 2).alignment = Alignment(horizontal="center")

        for col_idx, val in [(3, amount), (4, prev), (5, diff)]:
            c = ws.cell(row_num, col_idx, val)
            c.number_format = '#,##0'
            c.border = THIN

        pct_cell = ws.cell(row_num, 6, pct / 100)
        pct_cell.number_format = '+0.0%;-0.0%'
        pct_cell.border = THIN
        if pct > 0:
            pct_cell.font = Font(color="C0392B")
        elif pct < 0:
            pct_cell.font = Font(color="27AE60")

        ws.cell(row_num, 7, krw).number_format = '#,##0'
        ws.cell(row_num, 7).border = THIN

        # 교차 색상
        if i % 2 == 1:
            for col in range(1, 8):
                ws.cell(row_num, col).fill = GRAY

        total_krw += krw

    # ── 합계 ──
    total_row = 5 + len(rows)
    ws.cell(total_row, 1, "합계").font = Font(bold=True, size=11)
    ws.cell(total_row, 7, total_krw).number_format = '#,##0'
    ws.cell(total_row, 7).font = Font(bold=True, size=11, color="C0392B")
    for col in range(1, 8):
        ws.cell(total_row, col).fill = BLUE
        ws.cell(total_row, col).border = THIN

    # 열 너비
    widths = [14, 8, 16, 16, 14, 10, 18]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # ── 하단 서명란 ──
    sign_row = total_row + 3
    ws.cell(sign_row, 1, "작성").font = Font(bold=True)
    ws.cell(sign_row, 3, "검토").font = Font(bold=True)
    ws.cell(sign_row, 5, "승인").font = Font(bold=True)
    for col in [1, 3, 5]:
        ws.cell(sign_row + 1, col, "________________")

    wb.save(out_path)
    return len(rows), total_krw


if __name__ == "__main__":
    print("=== 데모 6: 보고서 템플릿 자동 채우기 ===")
    print()

    if not os.path.exists(DB_PATH):
        print(f"  [ERROR] DB not found: {DB_PATH}")
        exit(1)

    for month in ["2026-04", "2026-05", "2026-06"]:
        out = os.path.join(OUT, f"매출보고서_{month}.xlsx")
        cnt, krw = fill_monthly_report(DB_PATH, month, out)
        print(f"  {month}: {cnt}개 법인, 원화 합계 {krw:>15,}원 → {out}")

    print()
    print("  3개월치 보고서가 한 번에 생성되었습니다.")
    print("  → 매월 같은 양식에 숫자만 바꾸는 일을 자동화.")
    print("  → 전월 대비 증감률까지 자동 계산 + 색칠.")
    print("  → 서명란까지 포함된 인쇄용 양식.")
