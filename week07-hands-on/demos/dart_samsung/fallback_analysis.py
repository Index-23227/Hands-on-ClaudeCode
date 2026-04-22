"""
데모 B 폴백 — DART 삼성전자 재무비교 분석

동작:
1. .env에 DART_API_KEY 있으면 실제 DART API 호출
2. 실패 or 키 없음 → 내장 샘플 데이터 사용
3. 두 경우 모두 동일한 포맷의 엑셀 생성

실행:
  py week07-hands-on/demos/dart_samsung/fallback_analysis.py

출력:
  week07-hands-on/demos/dart_samsung/output/삼성전자_재무비교.xlsx
"""

from __future__ import annotations

import json
import ssl
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


HERE = Path(__file__).parent
OUT_DIR = HERE / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "삼성전자_재무비교.xlsx"

# DART 키 탐색 경로 (우선순위순)
# 1. week07-hands-on/prep/.env (이번 주차 전용)
# 2. week06-hands-on/.env (week06에서 이미 발급해 둔 키 재활용)
ENV_SEARCH_PATHS = [
    HERE.parent.parent / "prep" / ".env",
    HERE.parent.parent.parent / "week06-hands-on" / ".env",
]

SAMSUNG_CORP_CODE = "00126380"
YEARS = ["2024", "2025"]
REPRT_CODE = "11011"  # 사업보고서
FS_DIV = "CFS"        # 연결재무제표

# 주요 항목 account_id 매핑 (DART 표준)
TARGET_ACCOUNTS = {
    "ifrs-full_Revenue": "매출액",
    "dart_OperatingIncomeLoss": "영업이익",
    "ifrs-full_ProfitLoss": "당기순이익",
}

# 폴백용 샘플 (API 실패 시)
# 실제 공시값과 다를 수 있음. 시연 용도.
SAMPLE_DATA = {
    "2024": {
        "매출액":       258_935_000_000_000,
        "영업이익":      32_725_000_000_000,
        "당기순이익":    28_477_000_000_000,
    },
    "2025": {
        "매출액":       290_000_000_000_000,
        "영업이익":      42_300_000_000_000,
        "당기순이익":    35_400_000_000_000,
    },
}


def load_env() -> dict[str, str]:
    """여러 .env 후보 경로에서 병합 (앞쪽이 우선)"""
    env: dict[str, str] = {}
    for path in ENV_SEARCH_PATHS:
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            if k not in env:  # 앞쪽 파일 우선
                env[k] = v.strip().strip('"').strip("'")
    return env


def fetch_dart_year(api_key: str, year: str) -> dict[str, int] | None:
    """DART API에서 삼성전자 특정 연도 주요 재무 가져오기"""
    url = "https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json"
    params = {
        "crtfc_key": api_key,
        "corp_code": SAMSUNG_CORP_CODE,
        "bsns_year": year,
        "reprt_code": REPRT_CODE,
        "fs_div": FS_DIV,
    }
    full_url = url + "?" + urllib.parse.urlencode(params)
    ctx = ssl.create_default_context()
    try:
        req = urllib.request.Request(full_url, headers={"User-Agent": "week07-demo/1.0"})
        with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"    [경고] DART API 호출 실패 ({year}): {e}")
        return None

    if data.get("status") != "000":
        print(f"    [경고] DART API 응답 이상 ({year}): {data.get('status')} {data.get('message')}")
        return None

    result: dict[str, int] = {}
    for item in data.get("list", []):
        account_id = item.get("account_id", "")
        account_nm = item.get("account_nm", "")
        sj_div = item.get("sj_div", "")

        # 손익계산서(IS)만 대상으로 (재무상태표·현금흐름표 제외)
        if sj_div not in ("IS", "CIS"):
            continue

        amount_str = item.get("thstrm_amount", "").replace(",", "").strip()
        if not amount_str or amount_str == "-":
            continue
        try:
            amount = int(amount_str)
        except ValueError:
            continue
        if amount == 0:
            continue

        # account_id 우선 매칭
        kor_name = None
        if account_id in TARGET_ACCOUNTS:
            kor_name = TARGET_ACCOUNTS[account_id]
        elif account_nm in ("매출액", "영업이익", "당기순이익"):
            kor_name = account_nm

        if kor_name and kor_name not in result:
            result[kor_name] = amount

    if len(result) < 3:
        print(f"    [경고] {year}년 주요 3개 항목 중 일부 누락: {list(result.keys())}")
        return None

    return result


def get_financials() -> tuple[dict[str, dict[str, int]], str]:
    """실제 호출 시도 → 실패 시 샘플"""
    env = load_env()
    api_key = env.get("DART_API_KEY", "")
    if not api_key or "xxx" in api_key.lower():
        print("  [정보] DART_API_KEY 없음 → 샘플 데이터 사용")
        return SAMPLE_DATA, "샘플"

    print(f"  [시도] DART API 호출 (삼성전자 {', '.join(YEARS)})")
    results: dict[str, dict[str, int]] = {}
    for year in YEARS:
        data = fetch_dart_year(api_key, year)
        if data is None:
            print(f"  [실패] {year}년 데이터 가져오기 실패 → 전체 샘플로 전환")
            return SAMPLE_DATA, "샘플(API 실패)"
        results[year] = data
        print(f"    [OK] {year}: {', '.join(f'{k}={v/1e12:.1f}조' for k, v in data.items())}")

    return results, "DART 실시간"


def format_to_eok(amount: int) -> str:
    """원 → 조/억 표기"""
    if abs(amount) >= 1_000_000_000_000:
        return f"{amount / 1_000_000_000_000:.1f}조"
    if abs(amount) >= 100_000_000:
        return f"{amount / 100_000_000:.0f}억"
    return f"{amount:,}"


def build_xlsx(data: dict[str, dict[str, int]], data_source: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "재무비교"

    headers = ["항목", f"{YEARS[0]}년", f"{YEARS[1]}년", "증감액", "증감률"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 증가=빨강 / 감소=초록 (한국식)
    red_font = Font(color="C00000", bold=True)
    green_font = Font(color="00B050", bold=True)

    items = ["매출액", "영업이익", "당기순이익"]
    d1, d2 = data[YEARS[0]], data[YEARS[1]]

    for item in items:
        v1 = d1.get(item, 0)
        v2 = d2.get(item, 0)
        delta = v2 - v1
        rate = (delta / v1 * 100) if v1 else 0

        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=item)
        ws.cell(row=row_idx, column=2, value=format_to_eok(v1))
        ws.cell(row=row_idx, column=3, value=format_to_eok(v2))
        ws.cell(row=row_idx, column=4, value=format_to_eok(delta))
        rate_cell = ws.cell(row=row_idx, column=5, value=f"{rate:+.1f}%")

        # 증감률 색칠
        rate_cell.font = red_font if delta > 0 else (green_font if delta < 0 else Font())
        rate_cell.alignment = Alignment(horizontal="right")

    # 영업이익률·순이익률 추가
    ws.append([])  # 빈 행
    for label, numerator in [("영업이익률", "영업이익"), ("순이익률", "당기순이익")]:
        r1 = d1.get(numerator, 0) / d1["매출액"] * 100 if d1.get("매출액") else 0
        r2 = d2.get(numerator, 0) / d2["매출액"] * 100 if d2.get("매출액") else 0
        delta = r2 - r1
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=f"{r1:.1f}%")
        ws.cell(row=row_idx, column=3, value=f"{r2:.1f}%")
        ws.cell(row=row_idx, column=4, value=f"{delta:+.1f}%p")
        ws.cell(row=row_idx, column=5, value="")

    # 주목 포인트 3줄 요약 (폴백 텍스트)
    ws.append([])
    ws.append(["주목할 포인트"])
    summary_row = ws.max_row
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=13)

    rev_rate = (d2["매출액"] - d1["매출액"]) / d1["매출액"] * 100
    op_rate = (d2["영업이익"] - d1["영업이익"]) / d1["영업이익"] * 100
    op_margin_1 = d1["영업이익"] / d1["매출액"] * 100
    op_margin_2 = d2["영업이익"] / d2["매출액"] * 100

    notes = [
        f"1. 매출액이 전년 대비 {rev_rate:+.1f}% 성장. 반도체 메모리 사이클 반등 영향으로 해석 가능.",
        f"2. 영업이익이 {op_rate:+.1f}%로 매출 증가율보다 크게 개선 → 수익성 회복 구간 진입.",
        f"3. 영업이익률이 {op_margin_1:.1f}%에서 {op_margin_2:.1f}%로 개선. 단, 환율·메모리 가격 변동성 주시 필요.",
    ]
    for note in notes:
        ws.append([note])

    # 데이터 출처 표기
    ws.append([])
    ws.append([f"데이터 출처: {data_source}"])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="808080")

    # 열 너비
    widths = [20, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 2~4행 숫자 오른쪽 정렬
    for row_idx in range(2, 5):
        for col_idx in range(2, 6):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    wb.save(OUT_PATH)


def main() -> None:
    print("[시작] 삼성전자 재무비교 분석")
    data, source = get_financials()
    build_xlsx(data, source)
    print()
    print(f"[완료] {OUT_PATH}")
    print(f"       데이터 출처: {source}")


if __name__ == "__main__":
    main()
