"""
데모 F2 폴백 — 이상거래처 자동 탐지 + 담당자별 알림 메일 발송

동작:
1. data/shipments.xlsx 읽기
2. 이상 거래처 탐지 (품목 50개 초과 + 반품 0% or 반품 과다)
3. 담당자별 경각심 메일 본문 개인화 생성
4. 기본: 초안만 화면 출력 (발송 안 함)
5. --send 플래그: 강사 Gmail SMTP로 부계정들에게 실제 발송

실행:
  py week07-hands-on/demos/email_send/fallback_detect_and_send.py           # 초안만
  py week07-hands-on/demos/email_send/fallback_detect_and_send.py --send    # 실제 발송

안전장치:
- SUB1/SUB2/SUB3 placeholder는 .env의 SUB_ACCOUNTS 순서로 치환
- 수신자는 강사 부계정으로만 (외부 주소 불가)
- 전체 메일 본문 로그 → output/detection_report.txt
"""

from __future__ import annotations

import smtplib
import ssl
import sys
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from openpyxl import load_workbook

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


HERE = Path(__file__).parent
DATA_PATH = HERE / "data" / "shipments.xlsx"
OUT_DIR = HERE / "output"
OUT_DIR.mkdir(exist_ok=True)
REPORT_PATH = OUT_DIR / "detection_report.txt"
ENV_PATH = HERE.parent.parent / "prep" / ".env"


def load_env() -> dict[str, str]:
    if not ENV_PATH.exists():
        return {}
    env: dict[str, str] = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def load_shipments() -> list[dict]:
    if not DATA_PATH.exists():
        print(f"[에러] 데이터 없음: {DATA_PATH}")
        print(f"       prep/make_abnormal_shipments.py 먼저 실행")
        sys.exit(1)

    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb.active
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def detect_abnormal(rows: list[dict]) -> list[dict]:
    """거래처별 이상 등급 판정"""
    # 거래처별 집계
    by_cust: dict[str, dict] = defaultdict(lambda: {
        "items": defaultdict(lambda: [0, 0]),  # 품목코드 → [출고, 반품]
        "name": "",
        "contact": "",
        "email_label": "",
    })
    for r in rows:
        cust = r["거래처코드"]
        by_cust[cust]["name"] = r["거래처명"]
        by_cust[cust]["contact"] = r["담당자"]
        # email_label은 "SUB1@demo.local" → "SUB1"만 추출
        email_label = r["담당자이메일"].split("@")[0]
        by_cust[cust]["email_label"] = email_label
        it = by_cust[cust]["items"][r["품목코드"]]
        it[0] += r["출고수량"]
        it[1] += r["반품수량"]

    results = []
    for cust, info in by_cust.items():
        over_50 = [(code, q, rq) for code, (q, rq) in info["items"].items() if q > 50]
        if not over_50:
            continue

        total_q = sum(q for _, q, _ in over_50)
        total_rq = sum(rq for _, _, rq in over_50)
        return_rate = (total_rq / total_q * 100) if total_q else 0

        # 등급 판정
        if return_rate == 0 and len(over_50) >= 2:
            grade = "높음"
        elif return_rate == 0:
            grade = "주의"
        elif return_rate > 15:
            grade = "주의"
        else:
            grade = "정상범위"  # 알림 대상 아님

        results.append({
            "거래처코드": cust,
            "거래처명": info["name"],
            "담당자": info["contact"],
            "이메일라벨": info["email_label"],
            "이상품목수": len(over_50),
            "이상품목": over_50,
            "총출고": total_q,
            "총반품": total_rq,
            "반품률": return_rate,
            "등급": grade,
        })

    # 등급 우선순위 정렬
    grade_order = {"높음": 0, "주의": 1, "정상범위": 2}
    results.sort(key=lambda x: (grade_order[x["등급"]], x["거래처코드"]))
    return results


def build_email_body(detection: dict, sender_name: str = "재경팀") -> tuple[str, str]:
    """개인화된 메일 제목·본문"""
    subject = f"[가공매출 의심] {detection['거래처명']} 출고 패턴 점검 요청 (2026-03)"

    items_desc = "\n".join([
        f"  - {code}: 출고 {q}개, 반품 {rq}개"
        for code, q, rq in detection["이상품목"]
    ])

    if detection["등급"] == "높음":
        severity_note = (
            "반품 기록이 전혀 없으며 동일 품목이 다수 대량 출고되었습니다. "
            "가공매출 또는 신용/할인 리스크가 의심되는 패턴입니다."
        )
    elif detection["반품률"] == 0:
        severity_note = (
            "반품 기록이 전혀 없어 비정상 출고 가능성이 있습니다."
        )
    else:
        severity_note = (
            f"반품률이 {detection['반품률']:.0f}%로 비정상적으로 높습니다. "
            f"부도·허위출고 가능성을 검토해 주십시오."
        )

    body = f"""{detection['담당자']}님, 안녕하세요.

재경팀에서 2026년 3월 출고 데이터 모니터링 결과, 담당하신 {detection['거래처명']}({detection['거래처코드']})의
출고 패턴에서 다음과 같은 이상 징후가 감지되어 안내드립니다.

[감지 내용]
- 기준 초과 품목: {detection['이상품목수']}건 (동일 품목코드 50개 초과 출고)
{items_desc}
- 총 출고 수량: {detection['총출고']}개
- 총 반품 수량: {detection['총반품']}개
- 반품률: {detection['반품률']:.1f}%
- 이상 등급: [{detection['등급']}]

[해석]
{severity_note}

[요청]
영업팀 차원에서 해당 거래처 상태 및 출고 경위를 확인하신 뒤
금주 중 회신 부탁드립니다.

감사합니다.

{sender_name} 드림
"""
    return subject, body


def send_emails(send: bool, detections: list[dict], env: dict) -> None:
    # 알림 대상만 필터 (정상범위 제외)
    targets = [d for d in detections if d["등급"] in ("높음", "주의")]

    if not targets:
        print("[결과] 알림 대상 없음")
        return

    # .env 체크
    user = env.get("GMAIL_USER", "")
    password = env.get("GMAIL_APP_PASSWORD", "").replace(" ", "")
    subs_raw = env.get("SUB_ACCOUNTS", "")
    subs = [s.strip() for s in subs_raw.split(",") if s.strip()]

    # SUB1/SUB2/SUB3 → 실제 부계정 매핑
    def resolve_email(label: str) -> str:
        import re
        m = re.match(r"SUB(\d+)", label)
        if not m or not subs:
            return f"{label}@example.local"  # 폴백
        idx = (int(m.group(1)) - 1) % len(subs)
        return subs[idx]

    # 보고서 파일 작성
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write(f"=== 이상거래 탐지 리포트 ({2 * '-'} 메일 초안 포함) ===\n\n")
        for d in targets:
            subject, body = build_email_body(d)
            to_addr = resolve_email(d["이메일라벨"])
            f.write(f"[{d['등급']}] {d['거래처명']} → {to_addr}\n")
            f.write(f"제목: {subject}\n")
            f.write(f"본문:\n{body}\n")
            f.write("-" * 60 + "\n\n")

    print()
    print(f"[대상] 알림 발송 대상 {len(targets)}건")
    for d in targets:
        to_addr = resolve_email(d["이메일라벨"])
        print(f"  [{d['등급']}] {d['거래처명']:<14} → {to_addr}  (담당: {d['담당자']})")

    print()
    print(f"[보고서] {REPORT_PATH}")

    if not send:
        print()
        print("[모드] Dry-run — 실제 발송 안 함")
        print("      발송하려면: --send 플래그 추가")
        return

    # 실제 발송
    if not user or "your.email" in user:
        print("[에러] .env의 GMAIL_USER 미설정")
        sys.exit(1)
    if not password or "xxxx" in password:
        print("[에러] .env의 GMAIL_APP_PASSWORD 미설정")
        sys.exit(1)
    if not subs:
        print("[에러] .env의 SUB_ACCOUNTS 미설정 (부계정 없음)")
        sys.exit(1)

    print()
    print(f"[발송 시작] {user} → 부계정 {len(subs)}개 (순환 매핑)")

    ctx = ssl.create_default_context()
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ctx) as smtp:
            smtp.login(user, password)
            for d in targets:
                subject, body = build_email_body(d)
                to_addr = resolve_email(d["이메일라벨"])
                msg = MIMEMultipart()
                msg["From"] = user
                msg["To"] = to_addr
                msg["Subject"] = subject
                msg.attach(MIMEText(body, "plain", "utf-8"))
                smtp.sendmail(user, [to_addr], msg.as_string())
                print(f"  [SENT] {d['거래처명']:<14} → {to_addr}")
    except smtplib.SMTPAuthenticationError as e:
        print(f"[에러] Gmail 인증 실패: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"[에러] 발송 중 예외: {e}")
        sys.exit(1)

    print()
    print(f"[완료] {len(targets)}건 발송. 부계정 받은편지함에서 확인하세요.")


def main() -> None:
    send = "--send" in sys.argv

    print(f"[시작] 이상거래 탐지 → shipments.xlsx 로딩")
    rows = load_shipments()
    print(f"  총 거래 행: {len(rows)}행")
    print()

    detections = detect_abnormal(rows)

    print(f"[탐지 결과]")
    print(f"  {'거래처코드':<10} {'거래처명':<14} {'등급':<8} {'이상품목':<6} {'반품률':<8} {'담당자':<8}")
    print(f"  {'-'*10} {'-'*14} {'-'*8} {'-'*6} {'-'*8} {'-'*8}")
    for d in detections:
        print(f"  {d['거래처코드']:<10} {d['거래처명']:<14} {d['등급']:<8} "
              f"{d['이상품목수']:<6} {d['반품률']:>5.1f}%  {d['담당자']:<8}")

    env = load_env()
    send_emails(send, detections, env)


if __name__ == "__main__":
    main()
