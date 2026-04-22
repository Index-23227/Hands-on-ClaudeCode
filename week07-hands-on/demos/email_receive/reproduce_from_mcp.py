"""
F1 시연 결과 생성 스크립트 (MCP 흐름 재현용)

Gmail MCP(claude_ai_Gmail)로 8법인 메일 본문을 읽은 결과를 기반으로
week05 DB의 2026-03-31 월말 환율로 원화 환산 → 매출취합_2026-03.xlsx 생성.

시연 시 Claude는 이 스크립트를 직접 실행하지 않고, MCP 조회 결과를
openpyxl로 즉석 작성한다. 이 파일은 재현성 확보용.
"""

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HERE = Path(__file__).parent
DB_PATH = HERE.parents[2] / "week05-hands-on" / "data" / "sales.db"
OUT_DIR = HERE / "output"
OUT_DIR.mkdir(exist_ok=True)

# Gmail MCP 조회 결과 (본문 파싱)
REPORTS = [
    ("DE01", "독일법인",   "박서연", "EUR",            101_000, ""),
    ("US01", "미국법인",   "송예림", "USD",            118_000, "신규 거래처 2곳 추가"),
    ("VN01", "베트남법인", "강태우", "VND",  3_120_000_000,   ""),
    ("GB01", "영국법인",   "이민재", "GBP",             75_000, ""),
    ("IN01", "인도법인",   "정하늘", "INR",          8_800_000, ""),
    ("JP01", "일본법인",   "윤지우", "JPY",         14_900_000, ""),
    ("CN01", "중국법인",   "김준호", "CNY",            875_000, ""),
    ("TH01", "태국법인",   "최영훈", "THB",          4_100_000, ""),
]
EXPECTED_CORPS = {r[0] for r in REPORTS}


def load_rates() -> dict[str, float]:
    conn = sqlite3.connect(DB_PATH)
    rows = conn.execute(
        "SELECT currency, rate FROM exchange_rates WHERE rate_date = '2026-03-31'"
    ).fetchall()
    conn.close()
    return dict(rows)


def build() -> Path:
    rates = load_rates()

    wb = Workbook()
    ws = wb.active
    ws.title = "전체"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="D9D9D9")
    total_font = Font(bold=True)
    right = Alignment(horizontal="right")

    headers = ["법인코드", "법인명", "통화", "현지금액", "환율(KRW)", "원화환산(KRW)", "담당자", "비고"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    total_krw = 0
    sorted_reports = sorted(REPORTS, key=lambda r: r[1])  # 법인명 가나다순
    for code, corp, contact, currency, amount, note in sorted_reports:
        rate = rates[currency]
        krw = round(amount * rate)
        total_krw += krw
        ws.append([code, corp, currency, amount, rate, krw, contact, note])

    last = ws.max_row
    ws.append(["합계", "", "", "", "", total_krw, f"{len(REPORTS)}명", ""])
    for c in ws[last + 1]:
        c.fill = total_fill
        c.font = total_font

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[3].number_format = "#,##0"
        row[4].number_format = "#,##0.00"
        row[5].number_format = "#,##0"
        for col_idx in (3, 4, 5):
            row[col_idx].alignment = right

    widths = [10, 14, 6, 18, 12, 18, 10, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"

    # 시트2: 누락 리스트 (지금은 없지만 구조 유지)
    ws2 = wb.create_sheet("누락")
    ws2.append(["법인코드", "법인명", "담당자"])
    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
    received = {r[0] for r in REPORTS}
    missing = EXPECTED_CORPS - received
    if not missing:
        ws2.append(["— 누락 없음 —", "", ""])
    for code in sorted(missing):
        ws2.append([code, "", ""])

    out_path = OUT_DIR / "매출취합_2026-03.xlsx"
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    p = build()
    print(f"[OK] {p}")
