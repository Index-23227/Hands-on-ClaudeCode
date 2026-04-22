"""
데모 F1 폴백 — Gmail IMAP으로 법인 회신 메일 8통 자동 취합

MCP Gmail 설정이 안 됐거나 시연 중 실패했을 때 대체 실행.
prep/send_sample_inbox_emails.py 로 먼저 샘플 메일을 발송해둔 상태라야 함.

동작:
1. .env에서 GMAIL_USER, GMAIL_APP_PASSWORD 읽기
2. IMAP으로 최근 받은편지함 조회 → "매출 보고" 제목 필터
3. 각 메일에서 .xlsx 첨부 추출
4. 8법인 매출을 통합 엑셀로 저장 + KRW 환산 + 누락 법인 표시

실행:
  py week07-hands-on/demos/email_receive/fallback_imap.py

출력:
  week07-hands-on/demos/email_receive/output/매출취합_2026-03.xlsx
"""

from __future__ import annotations

import email
import imaplib
import sys
from email.header import decode_header
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")


HERE = Path(__file__).parent
OUT_DIR = HERE / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "매출취합_2026-03.xlsx"
ENV_PATH = HERE.parent.parent / "prep" / ".env"


# 기대하는 8법인 + 환율 (week05 DB 패턴 기준)
EXPECTED_LEGIONS = {
    "CN01": {"name": "중국법인", "contact": "김준호", "currency": "CNY", "rate": 186.0},
    "DE01": {"name": "독일법인", "contact": "박서연", "currency": "EUR", "rate": 1480.0},
    "GB01": {"name": "영국법인", "contact": "이민재", "currency": "GBP", "rate": 1720.0},
    "IN01": {"name": "인도법인", "contact": "정하늘", "currency": "INR", "rate": 16.0},
    "JP01": {"name": "일본법인", "contact": "윤지우", "currency": "JPY", "rate": 9.2},
    "TH01": {"name": "태국법인", "contact": "최영훈", "currency": "THB", "rate": 38.9},
    "US01": {"name": "미국법인", "contact": "송예림", "currency": "USD", "rate": 1350.0},
    "VN01": {"name": "베트남법인", "contact": "강태우", "currency": "VND", "rate": 0.056},
}


def load_env() -> dict[str, str]:
    if not ENV_PATH.exists():
        print(f"[에러] .env 없음: {ENV_PATH}")
        sys.exit(1)
    env: dict[str, str] = {}
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def decode_mime(raw: str) -> str:
    """MIME 인코딩된 문자열 디코드"""
    if not raw:
        return ""
    parts = decode_header(raw)
    result = []
    for text, enc in parts:
        if isinstance(text, bytes):
            try:
                result.append(text.decode(enc or "utf-8", errors="replace"))
            except (LookupError, TypeError):
                result.append(text.decode("utf-8", errors="replace"))
        else:
            result.append(text)
    return "".join(result)


def fetch_matching_emails(user: str, password: str) -> list[tuple[str, bytes]]:
    """받은편지함에서 '매출 보고' 제목 메일의 첨부 (파일명, 바이트) 리스트 반환

    Python imaplib은 검색 조건에 ASCII만 허용하므로,
    SINCE 날짜 기반으로 최근 30일 메일을 전부 가져온 뒤 Python에서 제목 필터링.
    """
    import datetime

    attachments: list[tuple[str, bytes]] = []

    with imaplib.IMAP4_SSL("imap.gmail.com", 993) as imap:
        imap.login(user, password)
        imap.select("INBOX")

        # 최근 30일 메일 가져오기 (ASCII-safe)
        since = (datetime.date.today() - datetime.timedelta(days=30)).strftime("%d-%b-%Y")
        status, msg_ids = imap.search(None, "SINCE", since)

        if status != "OK" or not msg_ids[0]:
            print(f"  [경고] 최근 30일 이내 메일 없음")
            return []

        ids = msg_ids[0].split()
        print(f"  [검색] 최근 30일 메일 {len(ids)}통 중 '매출 보고' 제목 필터링")

        matched = 0
        for msg_id in ids:
            status, data = imap.fetch(msg_id, "(RFC822)")
            if status != "OK":
                continue
            raw = data[0][1]
            msg = email.message_from_bytes(raw)
            subject = decode_mime(msg.get("Subject", ""))

            if "매출 보고" not in subject and "매출보고" not in subject:
                continue
            matched += 1

            for part in msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue
                # xlsx 첨부 파트만 대상 (Content-Type으로 판별)
                ctype = part.get_content_type()
                if ctype != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                    continue
                # filename 추출 3단계 fallback:
                # 1) 정석: get_filename()
                # 2) Content-Disposition raw 헤더가 RFC 2047로 인코딩된 경우 디코딩 → filename= 추출
                # 3) Content-Type의 name= 파라미터
                filename = decode_mime(part.get_filename() or "")
                if not filename.endswith(".xlsx"):
                    disp = part.get("Content-Disposition", "")
                    decoded_disp = decode_mime(disp)
                    # "attachment; filename=..." 또는 "attachment; filename*=utf-8''..." 파싱
                    import re as _re
                    m = _re.search(r'filename\*?=(?:"([^"]+)"|([^\s;]+))', decoded_disp)
                    if m:
                        filename = m.group(1) or m.group(2)
                if not filename.endswith(".xlsx"):
                    filename = decode_mime(part.get_param("name", "") or "")
                if not filename.endswith(".xlsx"):
                    continue
                payload = part.get_payload(decode=True)
                if payload:
                    attachments.append((filename, payload))
                    print(f"    [OK] {subject[:40]:<42} → {filename}")

        print(f"  [발견] 제목 매칭 {matched}건 / 첨부 {len(attachments)}건")

    return attachments


def parse_attachment(filename: str, data: bytes) -> tuple[str, str, float] | None:
    """첨부 엑셀에서 법인코드·통화·금액 추출"""
    # 파일명 패턴: 법인_{code}_{YYYY-MM}.xlsx
    import re
    m = re.match(r"법인_([A-Z]+\d+)_", filename)
    if not m:
        return None
    code = m.group(1)

    wb = load_workbook(BytesIO(data), data_only=True)
    ws = wb.active
    # 첫 데이터 행(2행) — 월|계정과목|통화|금액|비고
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        if len(row) >= 4 and row[3]:
            currency = str(row[2]).strip() if row[2] else ""
            amount = float(row[3])
            return (code, currency, amount)
    return None


def build_consolidated_xlsx(received: dict[str, tuple[str, float]]) -> None:
    """통합 엑셀 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "매출취합"

    headers = ["법인코드", "법인명", "통화", "현지 금액", "KRW 환산", "담당자", "상태"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for c_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=c_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    total_krw = 0.0
    missing: list[str] = []

    for code in sorted(EXPECTED_LEGIONS.keys()):
        info = EXPECTED_LEGIONS[code]
        if code in received:
            currency, amount = received[code]
            krw = amount * info["rate"]
            total_krw += krw
            ws.append([
                code, info["name"], currency, amount, krw,
                info["contact"], "수신 완료"
            ])
        else:
            missing.append(code)
            ws.append([
                code, info["name"], info["currency"], "",
                "", info["contact"], "미수신"
            ])
            # 미수신 행 빨간 배경
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=c_idx).fill = PatternFill(
                    start_color="FFE5E5", end_color="FFE5E5", fill_type="solid"
                )

    # 합계 행
    ws.append(["", "합계", "", "", total_krw, f"수신 {8 - len(missing)}/8", ""])
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_font = Font(bold=True)
    for c_idx in range(1, len(headers) + 1):
        c = ws.cell(row=ws.max_row, column=c_idx)
        c.fill = total_fill
        c.font = total_font

    # 숫자 서식
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (4, 5):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    widths = [10, 14, 8, 16, 16, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # 시트2: 누락 담당자
    if missing:
        ws2 = wb.create_sheet("누락 담당자")
        ws2.append(["법인코드", "법인명", "담당자", "상태", "재요청 메일 초안"])
        for c_idx in range(1, 6):
            c = ws2.cell(row=1, column=c_idx)
            c.fill = header_fill
            c.font = header_font

        for code in missing:
            info = EXPECTED_LEGIONS[code]
            draft = (f"{info['contact']}님, 안녕하세요.\n"
                     f"2026년 3월 매출 보고가 아직 회신되지 않았습니다. "
                     f"오늘 중 회신 부탁드립니다.\n감사합니다.")
            ws2.append([code, info["name"], info["contact"], "미수신", draft])

        widths2 = [10, 14, 10, 10, 60]
        for i, w in enumerate(widths2, start=1):
            ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    wb.save(OUT_PATH)
    return total_krw, missing


def main() -> None:
    env = load_env()
    user = env.get("GMAIL_USER", "")
    password = env.get("GMAIL_APP_PASSWORD", "").replace(" ", "")

    if not user or "your.email" in user:
        print("[에러] .env의 GMAIL_USER 설정 필요")
        sys.exit(1)
    if not password or "xxxx" in password:
        print("[에러] .env의 GMAIL_APP_PASSWORD 설정 필요")
        sys.exit(1)

    print(f"[시작] {user} IMAP 연결 → 매출 보고 메일 검색")

    try:
        attachments = fetch_matching_emails(user, password)
    except imaplib.IMAP4.error as e:
        print(f"[에러] IMAP 인증/연결 실패: {e}")
        sys.exit(1)

    if not attachments:
        print("[경고] 매칭 메일이 없습니다. prep/send_sample_inbox_emails.py 먼저 실행하세요.")
        sys.exit(1)

    # 첨부 파싱
    received: dict[str, tuple[str, float]] = {}
    for filename, data in attachments:
        parsed = parse_attachment(filename, data)
        if parsed:
            code, currency, amount = parsed
            received[code] = (currency, amount)

    print(f"  [파싱] {len(received)}/8 법인 매출 추출")

    # 통합 엑셀
    total_krw, missing = build_consolidated_xlsx(received)

    print()
    print(f"[완료] {OUT_PATH}")
    print(f"       수신 {len(received)}/8 법인, 합계 {total_krw:,.0f} 원")
    if missing:
        print(f"       누락: {', '.join(missing)}")


if __name__ == "__main__":
    main()
