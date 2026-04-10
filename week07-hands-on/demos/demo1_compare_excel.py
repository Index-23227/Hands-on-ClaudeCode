"""
데모 1: 두 엑셀 비교 — 이번 달 vs 저번 달 차이 자동 추출
"매달 눈으로 비교하던 걸 3초에 끝냅니다"
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "output")
os.makedirs(OUT, exist_ok=True)

# ── 1. 샘플 데이터 생성 ──
def create_sample():
    # 저번 달
    wb = Workbook()
    ws = wb.active
    ws.title = "매출"
    ws.append(["법인", "계정", "금액", "비고"])
    ws.append(["미국법인", "매출액", 125000, ""])
    ws.append(["일본법인", "매출액", 15800000, ""])
    ws.append(["중국법인", "매출액", 890000, "춘절 영향"])
    ws.append(["독일법인", "매출액", 98000, ""])
    ws.append(["베트남법인", "매출액", 3250000000, ""])
    ws.append(["인도법인", "매출액", 8500000, ""])
    ws.append(["영국법인", "매출액", 72000, ""])
    ws.append(["태국법인", "매출액", 4200000, ""])
    last = os.path.join(OUT, "3월_매출.xlsx")
    wb.save(last)

    # 이번 달 — 일부 금액 변경, 비고 추가
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "매출"
    ws2.append(["법인", "계정", "금액", "비고"])
    ws2.append(["미국법인", "매출액", 141000, ""])            # 변경
    ws2.append(["일본법인", "매출액", 15800000, ""])           # 동일
    ws2.append(["중국법인", "매출액", 960000, ""])             # 변경+비고삭제
    ws2.append(["독일법인", "매출액", 112000, "신규 거래처"])  # 변경+비고추가
    ws2.append(["베트남법인", "매출액", 3600000000, ""])       # 변경
    ws2.append(["인도법인", "매출액", 8500000, ""])            # 동일
    ws2.append(["영국법인", "매출액", 81000, ""])              # 변경
    ws2.append(["태국법인", "매출액", 4700000, ""])            # 변경
    this = os.path.join(OUT, "4월_매출.xlsx")
    wb2.save(this)
    return last, this


# ── 2. 비교 로직 ──
def compare(path_old, path_new):
    from openpyxl import load_workbook
    wb_old = load_workbook(path_old)
    wb_new = load_workbook(path_new)
    ws_old = wb_old.active
    ws_new = wb_new.active

    old_data = {}
    for i, row in enumerate(ws_old.iter_rows(values_only=True)):
        if i == 0:
            continue
        old_data[row[0]] = row  # key = 법인명

    diffs = []
    for i, row in enumerate(ws_new.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[0]
        old = old_data.get(name)
        if old is None:
            diffs.append({"법인": name, "변경": "신규 추가", "이전": "-", "이후": row[2]})
        elif old[2] != row[2]:
            change = row[2] - old[2]
            pct = (change / old[2] * 100) if old[2] else 0
            diffs.append({
                "법인": name,
                "변경": "금액 변경",
                "이전": f"{old[2]:,.0f}",
                "이후": f"{row[2]:,.0f}",
                "차이": f"{change:+,.0f}",
                "변동률": f"{pct:+.1f}%",
            })

    return diffs


# ── 3. 결과를 엑셀로 저장 ──
def save_diff_report(diffs, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "비교 결과"
    headers = ["법인", "변경", "이전 금액", "이후 금액", "차이", "변동률"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C3E50")

    red = Font(color="C0392B", bold=True)
    green = Font(color="27AE60", bold=True)

    for d in diffs:
        row = [d["법인"], d["변경"], d.get("이전", ""), d.get("이후", ""),
               d.get("차이", ""), d.get("변동률", "")]
        ws.append(row)
        # 증가면 빨강, 감소면 초록
        last_row = ws.max_row
        if d.get("차이", "").startswith("+"):
            ws.cell(last_row, 5).font = red
            ws.cell(last_row, 6).font = red
        elif d.get("차이", "").startswith("-"):
            ws.cell(last_row, 5).font = green
            ws.cell(last_row, 6).font = green

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    wb.save(out_path)


if __name__ == "__main__":
    print("=== 데모 1: 두 엑셀 비교 ===")
    print()
    old_path, new_path = create_sample()
    print(f"  3월 데이터: {old_path}")
    print(f"  4월 데이터: {new_path}")

    diffs = compare(old_path, new_path)
    print(f"\n  변경된 법인: {len(diffs)}개")
    print(f"  {'─'*60}")
    for d in diffs:
        print(f"  {d['법인']:10s} | {d.get('이전',''):>15s} → {d.get('이후',''):>15s} | {d.get('차이',''):>12s} ({d.get('변동률','')})")

    report = os.path.join(OUT, "비교결과_3월vs4월.xlsx")
    save_diff_report(diffs, report)
    print(f"\n  결과 저장: {report}")
    print("  → 색칠된 엑셀 열어보세요!")
